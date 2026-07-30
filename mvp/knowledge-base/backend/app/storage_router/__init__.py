"""
混合存储路由模块

根据文档类型自动选择最佳存储策略：
- 原样存储：合同、发票、合规文件（需原文溯源）
- RAG 向量化：手册、FAQ、文档（需语义检索）
- 结构化提取：表格、表单、数据（需字段查询）
- 知识图谱：关系型数据（人员、项目依赖）

Knowledge Agent 参照 rag-optimization-guide.md 做最终决策。
"""

import os
from typing import Optional
from datetime import datetime


# ── 存储策略定义 ──────────────────────────────────

STORAGE_STRATEGIES = {
    "rag": {
        "desc": "RAG 向量化存储",
        "suitable": ["产品手册", "FAQ", "技术文档", "会议纪要", "Markdown", "教程"],
        "not_suitable": ["合同", "发票", "表格数据"],
    },
    "original": {
        "desc": "原样存储 + 全文索引",
        "suitable": ["合同", "发票", "合规文件", "法律文书", "证书", "报告"],
        "not_suitable": ["FAQ", "短文本"],
    },
    "structured": {
        "desc": "结构化提取到 SQLite",
        "suitable": ["Excel", "CSV", "表格", "表单", "数据库导出", "价格表"],
        "not_suitable": ["长文本", "合同"],
    },
    "graph": {
        "desc": "知识图谱（Neo4j/JSON Graph）",
        "suitable": ["组织架构", "项目依赖", "供应商关系", "人员关系"],
        "not_suitable": ["纯文本", "表格"],
    },
}


# ── 文件类型 → 存储策略映射 ──────────────────────

FILE_TYPE_ROUTING = {
    # 文档类 → RAG
    ".md": "rag",
    ".markdown": "rag",
    ".txt": "rag",
    ".pdf": "rag",  # PDF 默认 RAG，但如果是合同/发票则 original

    # 表格类 → 结构化
    ".xlsx": "structured",
    ".xls": "structured",
    ".csv": "structured",
    ".tsv": "structured",

    # 图片类 → 多模态（OCR + RAG）
    ".jpg": "multimodal",
    ".jpeg": "multimodal",
    ".png": "multimodal",
    ".gif": "multimodal",
    ".webp": "multimodal",

    # 代码类 → RAG
    ".py": "rag",
    ".js": "rag",
    ".ts": "rag",
    ".java": "rag",
    ".go": "rag",

    # 数据类 → 结构化
    ".json": "structured",
    ".xml": "structured",
}


# ── 内容特征 → 策略调整 ──────────────────────────

CONTRACT_KEYWORDS = ["合同", "协议", "甲方", "乙方", "合同编号", "签署", "盖章", "invoice", "发票", "收据"]
TABLE_INDICATORS = ["sheet", "table", "row", "column", "单元格", "行", "列"]
RELATIONSHIP_INDICATORS = ["上级", "下级", "依赖", "关联", "负责人", "汇报", "parent", "child", "depends"]


def detect_content_type(text: str, filename: str = "") -> str:
    """
    分析内容特征，判断文档类型。
    返回: "contract" | "table" | "relationship" | "document" | "code" | "image"
    """
    if not text:
        text = ""

    text_lower = text.lower()

    # 合同/法律文件
    contract_score = sum(1 for kw in CONTRACT_KEYWORDS if kw in text_lower)
    if contract_score >= 2:
        return "contract"

    # 关系型数据
    relation_score = sum(1 for kw in RELATIONSHIP_INDICATORS if kw in text_lower)
    if relation_score >= 3:
        return "relationship"

    # 表格数据
    table_score = sum(1 for kw in TABLE_INDICATORS if kw in text_lower)
    ext = os.path.splitext(filename)[1].lower()
    if ext in (".xlsx", ".xls", ".csv", ".tsv") or table_score >= 2:
        return "table"

    # 代码
    code_extensions = {".py", ".js", ".ts", ".java", ".go", ".rs", ".cpp", ".c"}
    if ext in code_extensions:
        return "code"

    # 图片
    image_extensions = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
    if ext in image_extensions:
        return "image"

    # 默认：普通文档
    return "document"


def route_storage(filename: str, text: str = "", file_size: int = 0) -> dict:
    """
    混合存储路由：根据文件类型 + 内容特征选择最佳存储策略。

    返回:
    {
        "strategy": "rag" | "original" | "structured" | "graph" | "multimodal",
        "reason": str,
        "content_type": str,
        "actions": list[str],   # 需要执行的操作步骤
    }
    """
    content_type = detect_content_type(text, filename)
    ext = os.path.splitext(filename)[1].lower()

    # ── 策略决策 ──
    if content_type == "contract":
        strategy = "original"
        reason = "合同/法律文件，需原文溯源，原样存储 + 全文索引"
        actions = ["保存原始文件", "建立全文索引", "不做切割"]
    elif content_type == "table":
        strategy = "structured"
        reason = "表格数据，结构化提取到 SQLite，支持字段查询"
        actions = ["解析表格结构", "提取字段到 SQLite", "保留原始文件备份"]
    elif content_type == "relationship":
        strategy = "graph"
        reason = "关系型数据，适合知识图谱存储"
        actions = ["提取实体和关系", "构建图谱节点和边", "保留原文备份"]
    elif content_type == "image":
        strategy = "multimodal"
        reason = "图片文件，OCR 提取文字后走 RAG"
        actions = ["OCR 提取文字", "文字走 RAG 向量化", "保留原图"]
    elif content_type == "code":
        strategy = "rag"
        reason = "代码文件，按函数/类切割后 RAG"
        actions = ["按代码结构切割", "RAG 向量化", "保留源文件"]
    else:
        # 普通文档 → RAG
        strategy = "rag"
        reason = "普通文档，语义切割后 RAG 向量化"
        actions = ["语义切割", "RAG 向量化", "保留原文"]

    return {
        "strategy": strategy,
        "reason": reason,
        "content_type": content_type,
        "actions": actions,
        "file_extension": ext,
        "file_size": file_size,
    }


def get_strategy_info() -> dict:
    """获取所有存储策略信息"""
    return {
        "strategies": STORAGE_STRATEGIES,
        "file_routing": FILE_TYPE_ROUTING,
        "total_strategies": len(STORAGE_STRATEGIES) + 1,  # +1 for multimodal
    }


# ── 策略执行 ──────────────────────────────────────

def execute_strategy(route_result: dict, filepath: str, text: str = "") -> dict:
    """
    根据路由结果执行实际操作。

    Args:
        route_result: route_storage() 的返回值
        filepath: 已保存的临时文件路径
        text: 已提取的文本内容（可选）

    Returns:
        {"status": "ok", "details": {...}} 或 {"status": "error", "message": ...}
    """
    strategy = route_result["strategy"]
    filename = os.path.basename(filepath)
    results = {"strategy": strategy, "actions_performed": []}

    try:
        if strategy == "original":
            # 原样保存 + 全文索引
            results.update(_execute_original(filepath, filename))
            results["actions_performed"].append("fulltext_index")

        elif strategy == "structured":
            # 结构化提取到 SQLite
            results.update(_execute_structured(filepath, filename))
            results["actions_performed"].append("structured_import")

        elif strategy == "rag":
            # 向量化入库（已有文本时直接入库）
            if text:
                results.update(_execute_rag(filepath, filename, text))
            else:
                results.update({"status": "skipped", "message": "无文本内容，跳过 RAG 索引"})
            results["actions_performed"].append("rag_index")

        elif strategy == "multimodal":
            # OCR + RAG
            results.update(_execute_multimodal(filepath, filename, text))
            results["actions_performed"].append("ocr_extraction")

        elif strategy == "graph":
            # 知识图谱
            results.update(_execute_graph(filepath, filename, text))
            results["actions_performed"].append("graph_build")

        else:
            results.update({"status": "error", "message": f"未知策略: {strategy}"})

    except Exception as e:
        results["status"] = "error"
        results["message"] = str(e)

    return results


def _execute_original(filepath: str, filename: str) -> dict:
    """原样保存 + 全文索引（简单版：写入 ChromaDB 整篇不切割）"""
    from ..store import add_documents

    # 读取文件全文
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        full_text = f.read()

    # 作为单个 chunk 入库
    add_documents([{"text": full_text, "metadata": {
        "source": filename,
        "storage": "original",
        "filepath": filepath,
    }}])

    return {
        "status": "ok",
        "message": f"原样保存 {filename}，已建立全文索引",
        "chars": len(full_text),
    }


def _execute_structured(filepath: str, filename: str) -> dict:
    """结构化提取：Excel/CSV → SQLite"""
    import sqlite3

    ext = os.path.splitext(filename)[1].lower()
    db_path = os.path.join(os.path.dirname(filepath), "..", "structured_data.db")

    if ext in (".csv", ".tsv"):
        # CSV 解析
        import csv
        delimiter = "\t" if ext == ".tsv" else ","
        conn = sqlite3.connect(db_path)
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, None)
            if not headers:
                return {"status": "error", "message": "CSV 文件无表头"}

            table_name = _safe_table_name(filename)
            conn.execute(f"DROP TABLE IF EXISTS [{table_name}]")
            cols = ", ".join(f'"{h}" TEXT' for h in headers)
            conn.execute(f"CREATE TABLE [{table_name}] ({cols})")

            placeholders = ", ".join("?" * len(headers))
            rows = [tuple(row[:len(headers)]) for row in reader if any(row)]
            conn.executemany(f"INSERT INTO [{table_name}] VALUES ({placeholders})", rows)
            conn.commit()
            conn.close()
            return {"status": "ok", "message": f"CSV → SQLite 表 [{table_name}]", "rows": len(rows)}

    elif ext in (".xlsx", ".xls"):
        # Excel 解析
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            tables = {}
            conn = sqlite3.connect(db_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
                table_name = _safe_table_name(f"{filename}_{sheet_name}")
                conn.execute(f"DROP TABLE IF EXISTS [{table_name}]")
                cols = ", ".join(f'"{h}" TEXT' for h in headers)
                conn.execute(f"CREATE TABLE [{table_name}] ({cols})")
                data_rows = [tuple(r[:len(headers)]) for r in rows[1:] if any(v for v in r if v is not None)]
                placeholders = ", ".join("?" * len(headers))
                conn.executemany(f"INSERT INTO [{table_name}] VALUES ({placeholders})", data_rows)
                tables[sheet_name] = len(data_rows)
            conn.commit()
            conn.close()
            wb.close()
            return {"status": "ok", "message": f"Excel → SQLite", "sheets": tables}
        except ImportError:
            return {"status": "error", "message": "缺少 openpyxl 库，无法解析 Excel"}
    else:
        return {"status": "error", "message": f"不支持的结构化文件类型: {ext}"}


def _execute_rag(filepath: str, filename: str, text: str) -> dict:
    """文本走 RAG 向量化"""
    from ..chunk import chunk_text
    from ..store import add_documents

    chunks = chunk_text(text, metadata={
        "source": filename,
        "storage": "rag",
        "filepath": filepath,
    })
    if chunks:
        count = add_documents(chunks)
        return {"status": "ok", "message": f"RAG 索引完成", "chunks": count}
    return {"status": "error", "message": "文本切割失败"}


def _execute_multimodal(filepath: str, filename: str, text: str = "") -> dict:
    """多模态：OCR 提取文字 + RAG"""
    ocr_text = ""

    # 尝试 OCR（如果安装了 PIL + pytesseract）
    ext = os.path.splitext(filename)[1].lower()
    if ext in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"):
        try:
            from PIL import Image
            img = Image.open(filepath)
            try:
                import pytesseract
                ocr_text = pytesseract.image_to_string(img)
            except ImportError:
                ocr_text = f"[图片文件: {filename}，OCR 库未安装，仅保留元数据]"
            img.close()
        except ImportError:
            ocr_text = f"[图片文件: {filename}]"

    # 即使没 OCR，也将图片元数据入库
    if not ocr_text.strip():
        ocr_text = text or f"[图片文件: {filename}]"

    return _execute_rag(filepath, filename, ocr_text)


def _execute_graph(filepath: str, filename: str, text: str = "") -> dict:
    """知识图谱：简单版 JSON 保存"""
    import json as json_module
    graph_dir = os.path.join(os.path.dirname(filepath), "..", "graph_data")
    os.makedirs(graph_dir, exist_ok=True)

    graph_file = os.path.join(graph_dir, f"{os.path.splitext(filename)[0]}.json")
    data = {
        "source": filename,
        "text": text[:5000],
        "entities": [],  # 后续由 Knowledge Agent 提取
        "relations": [],
        "created_at": datetime.now().isoformat(),
    }
    with open(graph_file, "w", encoding="utf-8") as f:
        json_module.dump(data, f, ensure_ascii=False, indent=2)

    # 同时做 RAG 方便搜索
    return _execute_rag(filepath, filename, text)


def _safe_table_name(filename: str) -> str:
    """安全表名：仅保留字母数字下划线"""
    import re
    name = os.path.splitext(os.path.basename(filename))[0]
    name = re.sub(r'[^a-zA-Z0-9_\u4e00-\u9fff]', '_', name)
    return f"data_{name[:50]}"
